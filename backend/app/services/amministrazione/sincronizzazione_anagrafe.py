"""
Servizio per sincronizzazione anagrafe nazionale
Processa file .gz contenente dati anagrafe e crea partite animali.

Senza pandas: usa BeautifulSoup+lxml per HTML, openpyxl per Excel, csv per CSV
per ridurre uso memoria (~300MB risparmiati).
"""
import csv
import gzip
import io
import os
import re
import tempfile
from typing import List, Dict, Optional, Tuple, Union, Any
from datetime import datetime, date
from decimal import Decimal
from collections import defaultdict

from bs4 import BeautifulSoup
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.amministrazione.partita_animale import (
    PartitaAnimale,
    TipoPartita,
    ModalitaGestionePartita,
)
from app.models.allevamento.sede import Sede
from app.models.allevamento.azienda import Azienda


def _cell_text(cell) -> str:
    """Estrae testo da cella HTML (th/td), gestisce colspan/rowspan semplificato."""
    if cell is None:
        return ""
    text = cell.get_text(separator=" ", strip=True) if hasattr(cell, "get_text") else (cell.text or "")
    return (text or "").strip()


def parse_date(date_str: Optional[str]) -> Optional[date]:
    """Converte stringa data in formato dd/mm/yyyy a date object"""
    if not date_str or date_str.strip() == '':
        return None
    try:
        # Prova vari formati
        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except ValueError:
                continue
        return None
    except Exception:
        return None


def extract_table_from_html(html_content: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Estrae tabella da HTML. Restituisce (nomi_colonne, lista di righe come dict)."""
    return extract_table_from_stream(io.StringIO(html_content))


def extract_table_from_stream(stream) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Estrae tabella da stream HTML con BeautifulSoup+lxml (senza pandas).
    Restituisce (columns, rows) dove rows è lista di dict colonna -> valore.
    """
    try:
        content = stream.read() if hasattr(stream, "read") else stream
        if isinstance(content, bytes):
            content = content.decode("utf-8", errors="ignore")
    except Exception as e:
        raise ValueError(f"Errore lettura stream: {e}")
    try:
        soup = BeautifulSoup(content, "lxml")
    except Exception:
        soup = BeautifulSoup(content, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        raise ValueError("Nessuna tabella trovata nell'HTML")
    best_columns, best_rows = [], []
    for table in tables:
        thead = table.find("thead")
        tbody = table.find("tbody") or table
        header_cells = []
        if thead:
            header_row = thead.find("tr")
            if header_row:
                header_cells = header_row.find_all(["th", "td"])
        if not header_cells:
            first_tr = table.find("tr")
            if first_tr:
                header_cells = first_tr.find_all(["th", "td"])
        if not header_cells:
            continue
        columns = [normalize_column_name(_cell_text(c)) or f"col_{i}" for i, c in enumerate(header_cells)]
        rows = []
        trs = tbody.find_all("tr") if tbody != table else table.find_all("tr")[1:]
        for tr in trs:
            cells = tr.find_all(["td", "th"])
            if len(cells) < len(columns):
                row = {columns[i]: _cell_text(cells[i]) if i < len(cells) else "" for i in range(len(columns))}
            else:
                row = {columns[i]: _cell_text(cells[i]) for i in range(len(columns))}
            rows.append(row)
        if len(rows) > len(best_rows):
            best_columns, best_rows = columns, rows
    if not best_rows:
        raise ValueError("Nessuna riga dati trovata nelle tabelle HTML")
    return (best_columns, best_rows)


def _excel_to_rows(gz_file) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Legge Excel da file-like e restituisce (columns, rows)."""
    from openpyxl import load_workbook
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
        tmp.write(gz_file.read())
        tmp.flush()
        wb = load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise ValueError("File Excel vuoto")
    header = [str(c).strip().upper().replace(" ", "_") if c else f"col_{i}" for i, c in enumerate(rows_iter[0])]
    columns = [normalize_column_name(h) for h in header]
    rows = [dict(zip(columns, (str(v).strip() if v is not None else "" for v in row))) for row in rows_iter[1:]]
    return (columns, rows)


def _csv_to_rows(gz_file) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Legge CSV (tab-separated) da file-like e restituisce (columns, rows)."""
    text = gz_file.read().decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text), delimiter="\t")
    lines = list(reader)
    if not lines:
        raise ValueError("File CSV vuoto")
    header = lines[0]
    columns = [normalize_column_name(h) for h in header]
    rows = []
    for line in lines[1:]:
        row = {}
        for i, col in enumerate(columns):
            row[col] = (line[i].strip() if i < len(line) else "")
        rows.append(row)
    return (columns, rows)


def process_gz_file(gz_content: Union[bytes, str, io.BytesIO]) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Decomprime file .gz e estrae dati anagrafe.
    Restituisce (nomi_colonne, lista di righe come dict).
    Senza pandas: HTML con BeautifulSoup, Excel con openpyxl, CSV con modulo csv.
    """
    if isinstance(gz_content, str):
        gz_buffer = open(gz_content, "rb")
        should_close = True
        file_size = os.path.getsize(gz_content)
    elif isinstance(gz_content, bytes):
        gz_buffer = io.BytesIO(gz_content)
        should_close = True
        file_size = len(gz_content)
    else:
        gz_buffer = gz_content
        should_close = False
        cur = gz_buffer.tell()
        gz_buffer.seek(0, io.SEEK_END)
        file_size = gz_buffer.tell()
        gz_buffer.seek(cur)
    MAX_FILE_SIZE = 10 * 1024 * 1024
    if file_size > MAX_FILE_SIZE:
        if should_close and hasattr(gz_buffer, "close"):
            gz_buffer.close()
        raise ValueError(f"File troppo grande. Dimensione massima: {MAX_FILE_SIZE / (1024*1024):.1f}MB")
    try:
        is_html = False
        try:
            with gzip.GzipFile(fileobj=gz_buffer, mode="rb") as gz_peek:
                sample = gz_peek.read(2048).decode("utf-8", errors="ignore").lower()
                if "<html" in sample or "<!doctype html" in sample or "<table" in sample:
                    is_html = True
        except Exception:
            pass
        gz_buffer.seek(0)
        with gzip.GzipFile(fileobj=gz_buffer, mode="rb") as gz_file:
            if is_html:
                try:
                    text_stream = io.TextIOWrapper(gz_file, encoding="utf-8", errors="ignore")
                    return extract_table_from_stream(text_stream)
                except Exception:
                    pass
        gz_buffer.seek(0)
        with gzip.GzipFile(fileobj=gz_buffer, mode="rb") as gz_file:
            if not is_html:
                try:
                    return _excel_to_rows(gz_file)
                except Exception:
                    pass
        gz_buffer.seek(0)
        with gzip.GzipFile(fileobj=gz_buffer, mode="rb") as gz_file:
            try:
                return _csv_to_rows(gz_file)
            except Exception as e:
                raise ValueError(f"Impossibile leggere il file. Formati: HTML, XLSX, CSV. Errore: {str(e)}")
    except Exception as e:
        if "troppo grande" in str(e).lower():
            raise e
        import gc
        gc.collect()
        raise ValueError(f"Errore elaborazione file: {str(e)}")
    finally:
        if should_close and hasattr(gz_buffer, "close"):
            gz_buffer.close()


def normalize_column_name(col: str) -> str:
    """Normalizza nome colonna rimuovendo spazi e caratteri speciali"""
    return re.sub(r'[^\w]', '_', col.strip().upper())


def _row_val(row: Dict[str, Any], key: str) -> Optional[str]:
    """Restituisce valore riga come stringa strippata o None se vuoto."""
    v = row.get(key)
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def group_ingressi(columns: List[str], rows: List[Dict[str, Any]], azienda_codice: str, azienda_id: int, db: Session, codici_stalla_gestiti: Optional[set] = None) -> List[Dict]:
    """
    Raggruppa ingressi per DATA_INGRESSO e CODICE_PROVENIENZA.
    Funziona con lista di dict (senza pandas).
    """
    ingressi = [r for r in rows if _row_val(r, "DATA_INGRESSO") and _row_val(r, "CODICE_PROVENIENZA")]
    if not ingressi:
        return []
    if codici_stalla_gestiti is None:
        from app.services.allevamento.codici_stalla_service import get_codici_stalla_gestiti
        codici_stalla_gestiti = get_codici_stalla_gestiti(db, azienda_id)
    grouped: Dict[Tuple[str, str], List[Dict]] = defaultdict(list)
    for r in ingressi:
        key = (_row_val(r, "DATA_INGRESSO") or "", _row_val(r, "CODICE_PROVENIENZA") or "")
        grouped[key].append(r)
    partite = []
    first_col = columns[0] if columns else None
    for (data_ingresso_str, codice_provenienza), group in grouped.items():
        data_ingresso = parse_date(data_ingresso_str)
        if not data_ingresso:
            continue
        capi_unici = list({_row_val(r, "CODICE_CAPO") for r in group if _row_val(r, "CODICE_CAPO")})
        numero_capi = len(capi_unici)
        animali_dati = {}
        for row in group:
            codice_capo = _row_val(row, "CODICE_CAPO")
            if not codice_capo:
                continue
            sesso = _row_val(row, "SESSO")
            if sesso and sesso.upper() in ("M", "F"):
                sesso = sesso.upper()
            else:
                sesso = None
            data_nascita_val = _row_val(row, "DATA_NASCITA")
            data_nascita = parse_date(data_nascita_val) if data_nascita_val else None
            data_modello_ingresso_val = _row_val(row, "DATA_MODELLO_INGRESSO")
            data_modello_ingresso = parse_date(data_modello_ingresso_val) if data_modello_ingresso_val else None
            data_estrazione_val = _row_val(row, "DATA_ESTRAZIONE_DATI")
            data_estrazione_dati = parse_date(data_estrazione_val) if data_estrazione_val else None
            animali_dati[codice_capo] = {
                "sesso": sesso,
                "razza": _row_val(row, "RAZZA"),
                "data_nascita": data_nascita,
                "codice_elettronico": _row_val(row, "CODICE_ELETTRONICO"),
                "codice_madre": _row_val(row, "CODICE_MADRE"),
                "identificativo_fiscale_provenienza": _row_val(row, "IDENTIFICATIVO_FISCALE_PROV_"),
                "specie_allevata_provenienza": _row_val(row, "SPECIE_ALLEVATA_PROV"),
                "data_modello_ingresso": data_modello_ingresso,
                "codice_assegnato_precedenza": _row_val(row, "CODICE_ASSEGNATO_IN_PRECEDENZA"),
                "data_estrazione_dati": data_estrazione_dati,
            }
        codice_provenienza_str = (codice_provenienza or "").strip().upper()
        is_interno = codice_provenienza_str in codici_stalla_gestiti
        motivi = [_row_val(r, "MOTIVO_INGRESSO") for r in group if _row_val(r, "MOTIVO_INGRESSO")]
        motivo = max(set(motivi), key=motivi.count) if motivi else None
        modelli = [_row_val(r, "NUMERO_MODELLO_INGRESSO") for r in group if _row_val(r, "NUMERO_MODELLO_INGRESSO")]
        numero_modello = max(set(modelli), key=modelli.count) if modelli else None
        codice_stalla_azienda_val = None
        if group and first_col:
            codice_stalla_azienda_val = _row_val(group[0], "AZIENDA_CODICE") or _row_val(group[0], first_col)
        partite.append({
            "tipo": TipoPartita.INGRESSO,
            "data": data_ingresso,
            "codice_stalla": (codice_provenienza or "").strip(),
            "numero_capi": numero_capi,
            "motivo": motivo.strip() if motivo else None,
            "numero_modello": numero_modello.strip() if numero_modello else None,
            "is_trasferimento_interno": is_interno,
            "codici_capi": capi_unici,
            "animali_dati": animali_dati,
            "azienda_codice": azienda_codice,
            "codice_stalla_azienda_from_file": codice_stalla_azienda_val,
        })
    return partite


def group_uscite(columns: List[str], rows: List[Dict[str, Any]], azienda_codice: str, azienda_id: int, db: Session, codici_stalla_gestiti: Optional[set] = None) -> List[Dict]:
    """
    Raggruppa uscite per DATA_USCITA_STALLA e CODICE_AZIENDA_DESTINAZIONE.
    Funziona con lista di dict (senza pandas).
    """
    uscite = [r for r in rows if _row_val(r, "DATA_USCITA_STALLA") and _row_val(r, "CODICE_AZIENDA_DESTINAZIONE")]
    if not uscite:
        return []
    if codici_stalla_gestiti is None:
        from app.services.allevamento.codici_stalla_service import get_codici_stalla_gestiti
        codici_stalla_gestiti = get_codici_stalla_gestiti(db, azienda_id)
    grouped: Dict[Tuple[str, str], List[Dict]] = defaultdict(list)
    for r in uscite:
        key = (_row_val(r, "DATA_USCITA_STALLA") or "", _row_val(r, "CODICE_AZIENDA_DESTINAZIONE") or "")
        grouped[key].append(r)
    partite = []
    first_col = columns[0] if columns else None
    for (data_uscita_str, codice_destinazione), group in grouped.items():
        data_uscita = parse_date(data_uscita_str)
        if not data_uscita:
            continue
        capi_unici = list({_row_val(r, "CODICE_CAPO") for r in group if _row_val(r, "CODICE_CAPO")})
        numero_capi = len(capi_unici)
        codice_destinazione_str = (codice_destinazione or "").strip().upper()
        is_interno = codice_destinazione_str in codici_stalla_gestiti
        motivi = [_row_val(r, "MOTIVO_USCITA") for r in group if _row_val(r, "MOTIVO_USCITA")]
        motivo = max(set(motivi), key=motivi.count) if motivi else None
        modelli = [_row_val(r, "NUMERO_MODELLO_USCITA") for r in group if _row_val(r, "NUMERO_MODELLO_USCITA")]
        numero_modello = max(set(modelli), key=modelli.count) if modelli else None
        animali_dati_uscita = {}
        for row in group:
            codice_capo = _row_val(row, "CODICE_CAPO")
            if not codice_capo:
                continue
            data_modello_uscita_val = _row_val(row, "DATA_MODELLO_USCITA")
            data_modello_uscita = parse_date(data_modello_uscita_val) if data_modello_uscita_val else None
            data_macellazione_val = _row_val(row, "DATA_MACELLAZIONE")
            data_macellazione = parse_date(data_macellazione_val) if data_macellazione_val else None
            data_provv_val = _row_val(row, "DATA_PROVVEDIMENTO")
            data_provvvedimento = parse_date(data_provv_val) if data_provv_val else None
            animali_dati_uscita[codice_capo] = {
                "data_modello_uscita": data_modello_uscita,
                "codice_fiera_destinazione": _row_val(row, "CODICE_FIERA_DESTINAZIONE"),
                "codice_stato_destinazione": _row_val(row, "CODICE_STATO_DESTINAZIONE"),
                "regione_macello_destinazione": _row_val(row, "REGIONE_MACELLO_DESTINAZIONE"),
                "codice_macello_destinazione": _row_val(row, "CODICE_MACELLO_DESTINAZIONE"),
                "codice_pascolo_destinazione": _row_val(row, "CODICE_PASCOLO_DESTINAZIONE"),
                "codice_circo_destinazione": _row_val(row, "CODICE_CIRCO_DESTINAZIONE"),
                "data_macellazione": data_macellazione,
                "abbattimento": _row_val(row, "ABBATTIMENTO__S_N_"),
                "data_provvvedimento": data_provvvedimento,
            }
        codice_stalla_azienda_val = None
        if group and first_col:
            codice_stalla_azienda_val = _row_val(group[0], "AZIENDA_CODICE") or _row_val(group[0], first_col)
        partite.append({
            "tipo": TipoPartita.USCITA,
            "data": data_uscita,
            "codice_stalla": (codice_destinazione or "").strip(),
            "numero_capi": numero_capi,
            "motivo": motivo.strip() if motivo else None,
            "numero_modello": numero_modello.strip() if numero_modello else None,
            "is_trasferimento_interno": is_interno,
            "codici_capi": capi_unici,
            "animali_dati": animali_dati_uscita,
            "azienda_codice": azienda_codice,
            "codice_stalla_azienda_from_file": codice_stalla_azienda_val,
        })
    return partite


def group_decessi(columns: List[str], rows: List[Dict[str, Any]], azienda_codice: str, azienda_id: int, db: Session, animali_esistenti_map: Optional[dict] = None) -> List[Dict]:
    """
    Raggruppa decessi (MOTIVO_USCITA = 'D' o '02') per data di uscita.
    Funziona con lista di dict (senza pandas).
    """
    def is_decesso(row: Dict) -> bool:
        m = _row_val(row, "MOTIVO_USCITA")
        if not m:
            return False
        m = m.strip().upper()
        return m in ("D", "02", "2")
    decessi = [
        r for r in rows
        if is_decesso(r) and _row_val(r, "DATA_USCITA_STALLA") and _row_val(r, "CODICE_CAPO")
    ]
    if not decessi:
        return []
    if animali_esistenti_map is None:
        from app.models.allevamento.animale import Animale
        codici_capi = list({_row_val(r, "CODICE_CAPO") for r in decessi if _row_val(r, "CODICE_CAPO")})
        if codici_capi:
            animali_tuples = db.query(Animale.id, Animale.auricolare).filter(
                Animale.auricolare.in_(codici_capi),
                Animale.azienda_id == azienda_id,
                Animale.deleted_at.is_(None),
            ).all()
            animali_esistenti_map = {aur: id_ for id_, aur in animali_tuples}
        else:
            animali_esistenti_map = {}
    first_col = columns[0] if columns else None
    gruppi_decessi: Dict[str, Dict] = {}
    for row in decessi:
        data_uscita_stalla_val = _row_val(row, "DATA_USCITA_STALLA")
        data_decesso = parse_date(data_uscita_stalla_val) if data_uscita_stalla_val else None
        if not data_decesso:
            continue
        codice_capo = _row_val(row, "CODICE_CAPO")
        animale_id = animali_esistenti_map.get(codice_capo) if codice_capo else None
        numero_modello = _row_val(row, "NUMERO_MODELLO_USCITA")
        codice_stalla_decesso = _row_val(row, "AZIENDA_CODICE") or (first_col and _row_val(row, first_col))
        data_modello_uscita_val = _row_val(row, "DATA_MODELLO_USCITA")
        data_modello_uscita = parse_date(data_modello_uscita_val) if data_modello_uscita_val else None
        data_macellazione_val = _row_val(row, "DATA_MACELLAZIONE")
        data_macellazione = parse_date(data_macellazione_val) if data_macellazione_val else None
        data_provv_val = _row_val(row, "DATA_PROVVEDIMENTO")
        data_provvvedimento = parse_date(data_provv_val) if data_provv_val else None
        data_key = data_decesso.isoformat() if isinstance(data_decesso, date) else str(data_decesso)
        if data_key not in gruppi_decessi:
            gruppi_decessi[data_key] = {
                "data_uscita": data_decesso,
                "numero_certificato_smaltimento": numero_modello,
                "codice_stalla_decesso": codice_stalla_decesso,
                "codici_capi": [],
                "animali_ids": [],
                "animali_esistenti": 0,
                "animali_non_esistenti": 0,
                "animali_dati": {},
            }
        g = gruppi_decessi[data_key]
        g["codici_capi"].append(codice_capo)
        if animale_id:
            g["animali_ids"].append(animale_id)
            g["animali_esistenti"] += 1
        else:
            g["animali_non_esistenti"] += 1
        g["animali_dati"][codice_capo] = {
            "codice_stalla_decesso": codice_stalla_decesso,
            "numero_modello_uscita": numero_modello,
            "data_modello_uscita": data_modello_uscita,
            "codice_fiera_destinazione": _row_val(row, "CODICE_FIERA_DESTINAZIONE"),
            "codice_stato_destinazione": _row_val(row, "CODICE_STATO_DESTINAZIONE"),
            "regione_macello_destinazione": _row_val(row, "REGIONE_MACELLO_DESTINAZIONE"),
            "codice_macello_destinazione": _row_val(row, "CODICE_MACELLO_DESTINAZIONE"),
            "codice_pascolo_destinazione": _row_val(row, "CODICE_PASCOLO_DESTINAZIONE"),
            "codice_circo_destinazione": _row_val(row, "CODICE_CIRCO_DESTINAZIONE"),
            "data_macellazione": data_macellazione,
            "abbattimento": _row_val(row, "ABBATTIMENTO__S_N_"),
            "data_provvvedimento": data_provvvedimento,
            "a_carico": True,
        }
    return list(gruppi_decessi.values())


def filter_existing_decessi(
    gruppi_decessi: List[Dict],
    azienda_id: int,
    db: Session
) -> List[Dict]:
    """
    Filtra i gruppi decessi già registrati nel database usando query batch per ottimizzazione.
    
    Un gruppo è considerato duplicato se esiste già un gruppo con stessa data_uscita e azienda_id.
    """
    if not gruppi_decessi:
        return []
    
    from app.models.allevamento.gruppo_decessi import GruppoDecessi
    from sqlalchemy import or_
    
    # Estrai tutte le date per query batch
    date_uscita = [
        gruppo.get('data_uscita')
        for gruppo in gruppi_decessi
        if gruppo.get('data_uscita')
    ]
    
    if not date_uscita:
        return gruppi_decessi
    
    # Query batch per trovare tutti i gruppi esistenti
    existing_gruppi = db.query(GruppoDecessi).filter(
        GruppoDecessi.azienda_id == azienda_id,
        GruppoDecessi.data_uscita.in_(date_uscita),
        GruppoDecessi.deleted_at.is_(None)
    ).all()
    
    # Crea un set di date esistenti per lookup veloce
    existing_dates = {g.data_uscita for g in existing_gruppi}
    
    # Filtra i gruppi
    gruppi_filtered = [
        gruppo
        for gruppo in gruppi_decessi
        if gruppo.get('data_uscita') not in existing_dates
    ]
    
    return gruppi_filtered


def filter_existing_partite(
    partite: List[Dict],
    azienda_id: int,
    db: Session
) -> List[Dict]:
    """
    Filtra le partite già esistenti nel database usando query batch per ottimizzazione.
    
    Una partita è considerata duplicata se esiste già una partita con:
    - stessa azienda_id
    - stesso tipo (ingresso/uscita)
    - stessa data
    - stesso codice_stalla
    - stesso numero_capi
    - E se è stata completamente processata (ha tutti gli animali collegati)
    
    Se una partita esiste ma non è stata completamente processata, viene inclusa
    per permettere di completarla.
    
    OTTIMIZZAZIONE: Usa query batch per evitare N+1 query problem.
    """
    if not partite:
        return []
    
    from app.models.amministrazione.partita_animale_animale import PartitaAnimaleAnimale
    from app.models.allevamento.animale import Animale
    from sqlalchemy import and_, or_, func, case, literal_column
    
    # Crea una lista di tuple (tipo, data, codice_stalla) per query batch
    partite_keys = [
        (partita['tipo'], partita['data'], partita['codice_stalla'])
        for partita in partite
    ]
    
    # Query batch per trovare tutte le partite esistenti
    conditions = []
    for tipo, data, codice_stalla in partite_keys:
        conditions.append(
            and_(
                PartitaAnimale.azienda_id == azienda_id,
                PartitaAnimale.tipo == tipo,
                PartitaAnimale.data == data,
                PartitaAnimale.codice_stalla == codice_stalla,
                PartitaAnimale.deleted_at.is_(None)
            )
        )
    
    if not conditions:
        return partite
    
    # Esegui query batch
    existing_partite = db.query(PartitaAnimale).filter(
        or_(*conditions)
    ).all()
    
    if not existing_partite:
        return partite
    
    # Crea un dict per lookup veloce: (tipo, data, codice_stalla) -> PartitaAnimale
    existing_map = {
        (p.tipo, p.data, p.codice_stalla): p
        for p in existing_partite
    }
    
    # Raccogli tutti gli ID delle partite esistenti per query batch degli animali
    existing_ids = [p.id for p in existing_partite]
    
    # Query batch per contare animali collegati per tutte le partite
    animali_collegati_map = {}
    if existing_ids:
        risultati = db.query(
            PartitaAnimaleAnimale.partita_animale_id,
            func.count(PartitaAnimaleAnimale.id).label('count')
        ).filter(
            PartitaAnimaleAnimale.partita_animale_id.in_(existing_ids)
        ).group_by(PartitaAnimaleAnimale.partita_animale_id).all()
        
        animali_collegati_map = {r.partita_animale_id: r.count for r in risultati}
    
    # OTTIMIZZAZIONE: Query batch per animali creati invece di query individuali
    # Raggruppa le partite per tipo e codice_stalla_azienda per fare meno query
    animali_creati_map = {}
    
    # Raccogli tutti i codici stalla e date per query batch
    codici_stalla_ingresso = set()
    codici_stalla_uscita = set()
    date_ingresso = set()
    date_uscita = set()
    
    for existing in existing_partite:
        if existing.codice_stalla_azienda:
            if existing.tipo.value == 'ingresso':
                codici_stalla_ingresso.add(existing.codice_stalla_azienda)
                date_ingresso.add(existing.data)
            else:
                codici_stalla_uscita.add(existing.codice_stalla_azienda)
                date_uscita.add(existing.data)
    
    # Query batch per ingressi (una sola query per tutti)
    if codici_stalla_ingresso and date_ingresso:
        ingresso_counts = db.query(
            Animale.codice_azienda_anagrafe,
            Animale.data_arrivo,
            func.count(Animale.id).label('count')
        ).filter(
            Animale.codice_azienda_anagrafe.in_(list(codici_stalla_ingresso)),
            Animale.data_arrivo.in_(list(date_ingresso)),
            Animale.deleted_at.is_(None)
        ).group_by(
            Animale.codice_azienda_anagrafe,
            Animale.data_arrivo
        ).all()
        
        # Mappa: (codice_stalla, data) -> count
        ingresso_map = {(r.codice_azienda_anagrafe, r.data_arrivo): r.count for r in ingresso_counts}
        
        # Associa i conteggi alle partite
        for existing in existing_partite:
            if existing.codice_stalla_azienda and existing.tipo.value == 'ingresso':
                key = (existing.codice_stalla_azienda, existing.data)
                animali_creati_map[existing.id] = ingresso_map.get(key, 0)
    
    # Query batch per uscite (una sola query per tutti)
    if codici_stalla_uscita and date_uscita:
        uscita_counts = db.query(
            Animale.codice_azienda_anagrafe,
            Animale.data_uscita,
            func.count(Animale.id).label('count')
        ).filter(
            Animale.codice_azienda_anagrafe.in_(list(codici_stalla_uscita)),
            Animale.data_uscita.in_(list(date_uscita)),
            Animale.deleted_at.is_(None)
        ).group_by(
            Animale.codice_azienda_anagrafe,
            Animale.data_uscita
        ).all()
        
        # Mappa: (codice_stalla, data) -> count
        uscita_map = {(r.codice_azienda_anagrafe, r.data_uscita): r.count for r in uscita_counts}
        
        # Associa i conteggi alle partite
        for existing in existing_partite:
            if existing.codice_stalla_azienda and existing.tipo.value == 'uscita':
                key = (existing.codice_stalla_azienda, existing.data)
                animali_creati_map[existing.id] = uscita_map.get(key, 0)
    
    # Filtra le partite
    partite_filtered = []
    for partita in partite:
        key = (partita['tipo'], partita['data'], partita['codice_stalla'])
        existing = existing_map.get(key)
        
        if existing:
            # Verifica se la partita è stata completamente processata
            animali_collegati = animali_collegati_map.get(existing.id, 0)
            animali_creati = animali_creati_map.get(existing.id, 0)
            animali_processati_reali = max(animali_collegati, animali_creati)
            numero_capi_esistente = existing.numero_capi
            
            # Se la partita esiste ma non è completamente processata, includila
            if animali_processati_reali < numero_capi_esistente:
                partite_filtered.append(partita)
        else:
            # Se non esiste, aggiungi alla lista filtrata
            partite_filtered.append(partita)
    
    return partite_filtered


def extract_azienda_codice_from_file(columns: List[str], rows: List[Dict[str, Any]]) -> Optional[str]:
    """
    Estrae AZIENDA_CODICE dalla prima colonna della prima riga del file.
    Questa colonna identifica per quale codice stalla viene richiesto l'import.
    Senza pandas: lavora su (columns, rows).
    """
    if not rows:
        return None
    first_row = rows[0]
    # Colonne già normalizzate dal chiamante / da process_gz_file
    if "AZIENDA_CODICE" in columns:
        val = first_row.get("AZIENDA_CODICE")
        if val is not None and str(val).strip():
            return str(val).strip().upper()
    if columns:
        val = first_row.get(columns[0])
        if val is not None and str(val).strip():
            return str(val).strip().upper()
    return None


def verify_codice_stalla_exists(codice_stalla: str, db: Session) -> Tuple[bool, Optional[int], Optional[int]]:
    """
    Verifica se un codice stalla esiste nel database.
    
    Returns:
        (exists, sede_id, azienda_id) se esiste, altrimenti (False, None, None)
    """
    from app.models.allevamento.sede import Sede
    
    codice_norm = str(codice_stalla).strip().upper() if codice_stalla else None
    sede = db.query(Sede).filter(
        Sede.codice_stalla.ilike(codice_norm),
        Sede.deleted_at.is_(None)
    ).first()
    
    if sede:
        return True, sede.id, sede.azienda_id
    return False, None, None


def process_anagrafe_file(
    gz_content: Union[bytes, str],
    azienda_id: int,
    azienda_codice: str,
    db: Session
) -> Tuple[List[Dict], List[Dict], List[Dict], Optional[str]]:
    """
    Processa file anagrafe .gz e restituisce liste di partite e decessi da creare
    Esclude automaticamente le partite e decessi già presenti nel database.
    
    Ottimizzato per ridurre l'uso della memoria.
    Accetta bytes o percorso file (str).
    """
    # Forza garbage collection all'inizio per assicurare massima memoria disponibile
    import gc
    gc.collect()
    
    # Estrai dati dal file (columns, rows) senza pandas
    try:
        columns, rows = process_gz_file(gz_content)
    except Exception as e:
        if isinstance(gz_content, bytes):
            del gz_content
        gc.collect()
        raise e
    
    if isinstance(gz_content, bytes):
        del gz_content
    gc.collect()
    
    columns = [normalize_column_name(c) for c in columns]
    
    MAX_ROWS = 25000
    if len(rows) > MAX_ROWS:
        raise ValueError(f"File contiene troppe righe ({len(rows)}). Dimensione massima consentita: {MAX_ROWS} righe. Suddividere il file o contattare il supporto.")
    
    codice_stalla_file = extract_azienda_codice_from_file(columns, rows)
    
    # Verifica se il codice stalla esiste
    if codice_stalla_file:
        sede_exists, sede_id, sede_azienda_id = verify_codice_stalla_exists(codice_stalla_file, db)
        if not sede_exists:
            # Il codice stalla non esiste, verrà gestito dall'endpoint per richiedere la creazione
            pass  # Continuiamo comunque il processamento per identificare le partite
    
    # Carica codici stalla gestiti una sola volta per ottimizzazione
    from app.services.allevamento.codici_stalla_service import get_codici_stalla_gestiti
    codici_stalla_gestiti = get_codici_stalla_gestiti(db, azienda_id)
    
    # Pre-carica animali esistenti per ottimizzare group_decessi
    # OTTIMIZZAZIONE: Carica solo ID e Auricolare, non interi oggetti ORM
    from app.models.allevamento.animale import Animale
    
    all_codici_capi = set()
    if "CODICE_CAPO" in columns:
        for r in rows:
            v = r.get("CODICE_CAPO")
            if v is not None and str(v).strip():
                all_codici_capi.add(str(v).strip())
    
    animali_esistenti_map = {}
    if all_codici_capi:
        # Usa yield_per per processare in batch se sono troppi, ma qui basta selezionare solo le colonne
        # Carica tuple (id, auricolare) invece di oggetti completi per risparmiare >80% memoria
        animali_tuples = db.query(Animale.id, Animale.auricolare).filter(
            Animale.auricolare.in_(list(all_codici_capi)),
            Animale.azienda_id == azienda_id,
            Animale.deleted_at.is_(None)
        ).all()
        
        animali_esistenti_map = {aur: id_ for id_, aur in animali_tuples}
    
    partite_ingresso = group_ingressi(columns, rows, azienda_codice, azienda_id, db, codici_stalla_gestiti)
    partite_uscita = group_uscite(columns, rows, azienda_codice, azienda_id, db, codici_stalla_gestiti)
    
    # Importa il service per i codici stalla
    from app.services.allevamento.codici_stalla_service import (
        determina_codice_stalla_azienda,
        is_codice_stalla_gestito,
        get_codice_stalla_default_ingresso,
        get_codice_stalla_default_uscita
    )
    
    # Aggiungi codice_stalla_azienda a tutte le partite (codice stalla dell'allevamento dell'utente)
    # Determina dinamicamente in base alla logica di business
    # Priorità: codice_stalla_azienda_from_file (estratto dalla prima colonna) > codice_stalla_file > logica dinamica
    
    for partita in partite_ingresso + partite_uscita:
        tipo = partita['tipo'].value
        is_trasferimento_interno = partita.get('is_trasferimento_interno', False)
        
        # Priorità 1: usa codice_stalla_azienda_from_file se disponibile (estratto dalla prima colonna del file)
        codice_stalla_azienda = partita.get('codice_stalla_azienda_from_file')
        
        # Priorità 2: se non disponibile, usa codice_stalla_file
        if not codice_stalla_azienda:
            # Per ingressi da esterni: codice_stalla_file è la destinazione (sede per cui viene fatto l'import)
            # Per ingressi interni: determina dinamicamente la destinazione
            # Per uscite: codice_stalla_file è l'origine (sede per cui viene fatto l'import)
            if tipo == 'ingresso' and not is_trasferimento_interno:
                # Ingresso da esterno: usa codice_stalla_file come destinazione
                codice_stalla_azienda = codice_stalla_file
            elif tipo == 'uscita' and not is_trasferimento_interno:
                # Uscita verso esterno: usa codice_stalla_file come origine
                codice_stalla_azienda = codice_stalla_file
            else:
                # Trasferimenti interni: determina dinamicamente
                codice_stalla_azienda = determina_codice_stalla_azienda(
                    codice_stalla_provenienza=partita['codice_stalla'],
                    tipo=tipo,
                    is_trasferimento_interno=is_trasferimento_interno,
                    db=db,
                    azienda_id=azienda_id
                )
                
                # Fallback solo se determina_codice_stalla_azienda restituisce None
                if not codice_stalla_azienda:
                    if tipo == 'ingresso':
                        codice_stalla_azienda = get_codice_stalla_default_ingresso(db, azienda_id)
                    elif tipo == 'uscita':
                        codice_stalla_azienda = get_codice_stalla_default_uscita(db, azienda_id)
                    
                    # Ultimo fallback: codice_stalla_file
                    if not codice_stalla_azienda:
                        codice_stalla_azienda = codice_stalla_file
        
        partita['codice_stalla_azienda'] = codice_stalla_azienda
    
    decessi = group_decessi(columns, rows, azienda_codice, azienda_id, db, animali_esistenti_map)
    
    del rows
    gc.collect()
    
    # Aggiungi azienda_id a tutti i gruppi decessi
    for gruppo in decessi:
        gruppo['azienda_id'] = azienda_id
    
    # Filtra le partite già esistenti nel database
    partite_ingresso = filter_existing_partite(partite_ingresso, azienda_id, db)
    partite_uscita = filter_existing_partite(partite_uscita, azienda_id, db)
    
    # Filtra i decessi già esistenti nel database
    decessi = filter_existing_decessi(decessi, azienda_id, db)
    
    return partite_ingresso, partite_uscita, decessi, codice_stalla_file


def create_partite_from_groups(
    partite_data: List[Dict],
    azienda_id: int,
    file_anagrafe_origine: str,
    db: Session
) -> List[PartitaAnimale]:
    """
    Crea record PartitaAnimale nel database a partire dai dati raggruppati
    """
    created_partite = []
    
    for partita_data in partite_data:
        # Genera numero_partita univoco se non fornito
        numero_partita = partita_data.get('numero_partita')
        if not numero_partita:
            tipo_prefix = "ING" if partita_data['tipo'] == TipoPartita.INGRESSO else "USC"
            data_str = partita_data['data'].strftime('%Y%m%d')
            codice_stalla = partita_data['codice_stalla']
            
            # Conta quante partite esistono già per questo giorno e tipo
            count = db.query(func.count(PartitaAnimale.id)).filter(
                PartitaAnimale.data == partita_data['data'],
                PartitaAnimale.tipo == partita_data['tipo'],
                PartitaAnimale.deleted_at.is_(None)
            ).scalar() or 0
            
            # Genera numero progressivo a 3 cifre (001, 002, ecc.)
            numero_progressivo = f"{(count + 1):03d}"
            numero_partita = f"{tipo_prefix}-{data_str}-{codice_stalla}-{numero_progressivo}"
        
        # Crea partita
        partita = PartitaAnimale(
            azienda_id=azienda_id,
            tipo=partita_data['tipo'],
            data=partita_data['data'],
            numero_partita=numero_partita,
            codice_stalla=partita_data['codice_stalla'],
            nome_stalla=partita_data.get('nome_stalla'),
            numero_capi=partita_data['numero_capi'],
            peso_totale=partita_data.get('peso_totale'),
            peso_medio=partita_data.get('peso_medio'),
            modalita_gestione=ModalitaGestionePartita.PROPRIETA.value,
            motivo=partita_data.get('motivo'),
            numero_modello=partita_data.get('numero_modello'),
            is_trasferimento_interno=partita_data.get('is_trasferimento_interno', False),
            file_anagrafe_origine=file_anagrafe_origine,
            data_importazione=datetime.utcnow(),
            note=partita_data.get('note'),
        )
        
        db.add(partita)
        created_partite.append(partita)
    
    db.commit()
    
    # Refresh per ottenere gli ID
    for partita in created_partite:
        db.refresh(partita)
    
    return created_partite

