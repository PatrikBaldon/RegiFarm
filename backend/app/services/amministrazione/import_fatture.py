"""
Servizio per importare fatture da file Excel (.xlsx) e XML FatturaPA

NOTA: pandas viene importato LAZY (solo quando necessario) per risparmiare RAM all'avvio.
Questo è importante perché pandas occupa ~50-100MB di memoria.
"""
# LAZY IMPORT: pandas viene importato nelle funzioni che lo usano
import xml.etree.ElementTree as ET
from typing import List, Dict, Optional, Tuple
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import Session
from sqlalchemy import and_
import os
import glob
import zipfile
import tempfile
import shutil
import math

# from app.models.amministrazione.fattura_emessa import FatturaEmessa, StatoPagamentoFatturaEmessa  # DEPRECATO: usa FatturaAmministrazione
from app.models.amministrazione.fattura_amministrazione import FatturaAmministrazione, TipoFattura, StatoPagamento
from app.models.amministrazione.fornitore import Fornitore
from app.models.terreni.terreno import Terreno
from app.models.allevamento.azienda import Azienda


def _is_nan(value) -> bool:
    """Verifica se un valore è NaN senza dipendere da pandas"""
    if value is None:
        return True
    # Controlla per float NaN
    if isinstance(value, float) and math.isnan(value):
        return True
    # Controlla per stringhe vuote
    if isinstance(value, str) and value.strip() == '':
        return True
    return False


def parse_date(value) -> Optional[date]:
    """Converte un valore in date"""
    if _is_nan(value) or value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            # Prova vari formati di data
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
            return None
        except:
            return None
    return None


def parse_decimal(value) -> Decimal:
    """Converte un valore in Decimal"""
    if _is_nan(value) or value is None:
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        try:
            # Rimuovi spazi e caratteri non numerici (tranne punto e virgola)
            cleaned = value.strip().replace(' ', '').replace(',', '.')
            return Decimal(cleaned)
        except (InvalidOperation, ValueError):
            return Decimal('0')
    return Decimal('0')


def parse_string(value) -> Optional[str]:
    """Converte un valore in stringa"""
    if _is_nan(value) or value is None:
        return None
    return str(value).strip() if str(value).strip() != '' else None


def find_fornitore(db: Session, nome: Optional[str], piva: Optional[str], cf: Optional[str]) -> Optional[int]:
    """Trova un fornitore per nome o P.IVA
    
    Nota: Il modello Fornitore non ha codice_fiscale, solo partita_iva.
    Il parametro cf viene ignorato.
    """
    if not nome and not piva:
        return None
    
    query = db.query(Fornitore).filter(Fornitore.deleted_at.is_(None))
    
    # Cerca per P.IVA (priorità)
    if piva:
        fornitore = query.filter(Fornitore.partita_iva == piva).first()
        if fornitore:
            return fornitore.id
    
    # Cerca per nome (se non trovato per P.IVA)
    if nome:
        fornitore = query.filter(Fornitore.nome.ilike(f'%{nome}%')).first()
        if fornitore:
            return fornitore.id
    
    # Se abbiamo CF ma non P.IVA, potrebbe essere che CF == P.IVA per alcune aziende
    # In tal caso, proviamo a cercare per P.IVA usando il CF
    if cf and not piva:
        fornitore = query.filter(Fornitore.partita_iva == cf).first()
        if fornitore:
            return fornitore.id
    
    return None


def find_terreno(db: Session, azienda_id: int, nome: Optional[str]) -> Optional[int]:
    """Trova un terreno per nome nell'azienda"""
    if not nome:
        return None
    
    terreno = db.query(Terreno).filter(
        and_(
            Terreno.azienda_id == azienda_id,
            Terreno.nome.ilike(f'%{nome}%'),
            Terreno.deleted_at.is_(None)
        )
    ).first()
    
    return terreno.id if terreno else None


def parse_stato_pagamento_emessa(value: str) -> StatoPagamento:
    """Converte una stringa in StatoPagamento (per fatture emesse, tipo=entrata)"""
    if not value:
        return StatoPagamento.DA_INCASSARE
    
    value_lower = str(value).strip().lower().replace(' ', '_')
    
    stato_map = {
        'da_incassare': StatoPagamento.DA_INCASSARE,
        'incassata': StatoPagamento.INCASSATA,
        'scaduta': StatoPagamento.SCADUTA,
        'parziale': StatoPagamento.PARZIALE,
        'annullata': StatoPagamento.ANNULLATA,
        'pagata': StatoPagamento.INCASSATA,
        'da_pagare': StatoPagamento.DA_INCASSARE,
    }
    
    return stato_map.get(value_lower, StatoPagamento.DA_INCASSARE)


def parse_stato_pagamento_amministrazione(value: str) -> StatoPagamento:
    """Converte una stringa in StatoPagamento"""
    if not value:
        return StatoPagamento.DA_PAGARE
    
    value_lower = str(value).strip().lower().replace(' ', '_')
    
    stato_map = {
        'da_pagare': StatoPagamento.DA_PAGARE,
        'pagata': StatoPagamento.PAGATA,
        'scaduta': StatoPagamento.SCADUTA,
        'parziale': StatoPagamento.PARZIALE,
    }
    
    return stato_map.get(value_lower, StatoPagamento.DA_PAGARE)


def parse_tipo_fattura(value: str) -> TipoFattura:
    """Converte una stringa in TipoFattura"""
    if not value:
        return TipoFattura.USCITA
    
    value_lower = str(value).strip().lower()
    
    if 'entrata' in value_lower or 'ricavo' in value_lower or 'incasso' in value_lower:
        return TipoFattura.ENTRATA
    else:
        return TipoFattura.USCITA


def import_fatture_emesse(
    db: Session,
    file_path: str,
    azienda_id: int,
    skip_duplicates: bool = True
) -> Dict[str, any]:
    """
    Importa fatture emesse da file Excel
    
    Args:
        db: Database session
        file_path: Percorso del file Excel
        azienda_id: ID azienda
        skip_duplicates: Se True, salta le fatture duplicate (stesso numero e data)
    
    Returns:
        Dict con statistiche dell'import
    """
    # Usa openpyxl direttamente invece di pandas per risparmiare ~300MB di RAM
    from openpyxl import load_workbook
    
    try:
        # Verifica che l'azienda esista
        azienda = db.query(Azienda).filter(Azienda.id == azienda_id).first()
        if not azienda:
            return {
                'success': False,
                'error': f'Azienda con ID {azienda_id} non trovata',
                'importate': 0,
                'errate': 0,
                'duplicate': 0
            }
        
        # Leggi il file Excel con openpyxl
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {'success': False, 'error': 'File Excel vuoto', 'importate': 0, 'errate': 0, 'duplicate': 0}
        
        # Prima riga = header, normalizza nomi colonne
        header = [str(col).strip().lower().replace(' ', '_') if col else f'col_{i}' for i, col in enumerate(rows[0])]
        data_rows = rows[1:]
        
        # Converti in lista di dizionari (simile a df.to_dict('records'))
        df_records = [dict(zip(header, row)) for row in data_rows]
        
        importate = 0
        errate = 0
        duplicate = 0
        errori = []
        
        # Mappa delle colonne possibili
        colonne_richieste = ['numero', 'data_fattura', 'importo_totale']
        colonne_opzionali = {
            'cliente_nome': ['cliente_nome', 'cliente', 'nome_cliente', 'ragione_sociale'],
            'cliente_piva': ['cliente_piva', 'piva', 'partita_iva', 'p.iva'],
            'cliente_cf': ['cliente_cf', 'cf', 'codice_fiscale'],
            'data_registrazione': ['data_registrazione', 'data_reg'],
            'importo_iva': ['importo_iva', 'iva'],
            'importo_netto': ['importo_netto', 'netto'],
            'importo_incassato': ['importo_incassato', 'incassato'],
            'stato_pagamento': ['stato_pagamento', 'stato'],
            'data_scadenza': ['data_scadenza', 'scadenza'],
            'data_incasso': ['data_incasso', 'data_incasso_effettivo'],
            'aliquota_iva': ['aliquota_iva', 'aliquota'],
            'categoria': ['categoria', 'tipo_vendita'],
            'terreno': ['terreno', 'terreno_nome'],
            'note': ['note', 'osservazioni']
        }
        
        # Verifica colonne richieste
        colonne_trovate = {}
        for col in colonne_richieste:
            if col in header:
                colonne_trovate[col] = col
            else:
                return {
                    'success': False,
                    'error': f'Colonna obbligatoria "{col}" non trovata nel file',
                    'importate': 0,
                    'errate': 0,
                    'duplicate': 0
                }
        
        # Trova colonne opzionali
        for campo, possibili_nomi in colonne_opzionali.items():
            for nome in possibili_nomi:
                if nome in header:
                    colonne_trovate[campo] = nome
                    break
        
        # Processa ogni riga (df_records è lista di dict)
        for idx, row in enumerate(df_records):
            try:
                # Estrai dati base (row è un dict, accesso con .get())
                numero = parse_string(row.get(colonne_trovate['numero']))
                data_fattura = parse_date(row.get(colonne_trovate['data_fattura']))
                importo_totale = parse_decimal(row.get(colonne_trovate['importo_totale']))
                
                if not numero or not data_fattura or importo_totale == 0:
                    errate += 1
                    errori.append(f'Riga {idx + 2}: Dati obbligatori mancanti o non validi')
                    continue
                
                # Verifica duplicati
                if skip_duplicates:
                    esistente = db.query(FatturaAmministrazione).filter(
                        and_(
                            FatturaAmministrazione.azienda_id == azienda_id,
                            FatturaAmministrazione.tipo == TipoFattura.ENTRATA,
                            FatturaAmministrazione.numero == numero,
                            FatturaAmministrazione.data_fattura == data_fattura,
                            FatturaAmministrazione.deleted_at.is_(None)
                        )
                    ).first()
                    
                    if esistente:
                        duplicate += 1
                        continue
                
                # Estrai dati opzionali
                cliente_nome = parse_string(row.get(colonne_trovate.get('cliente_nome', ''), ''))
                cliente_piva = parse_string(row.get(colonne_trovate.get('cliente_piva', ''), ''))
                cliente_cf = parse_string(row.get(colonne_trovate.get('cliente_cf', ''), ''))
                
                # Trova fornitore/cliente
                cliente_id = find_fornitore(db, cliente_nome, cliente_piva, cliente_cf)
                
                # Calcola importi
                importo_iva = parse_decimal(row.get(colonne_trovate.get('importo_iva', ''), 0))
                importo_netto = parse_decimal(row.get(colonne_trovate.get('importo_netto', ''), 0))
                if importo_netto == 0:
                    importo_netto = importo_totale - importo_iva
                
                importo_incassato = parse_decimal(row.get(colonne_trovate.get('importo_incassato', ''), 0))
                
                # Stato pagamento
                stato_str = parse_string(row.get(colonne_trovate.get('stato_pagamento', ''), ''))
                stato_pagamento = parse_stato_pagamento_emessa(stato_str) if stato_str else StatoPagamento.DA_INCASSARE
                
                # Date
                data_registrazione = parse_date(row.get(colonne_trovate.get('data_registrazione', ''), '')) or date.today()
                data_scadenza = parse_date(row.get(colonne_trovate.get('data_scadenza', ''), ''))
                data_incasso = parse_date(row.get(colonne_trovate.get('data_incasso', ''), ''))
                
                # Aliquota IVA
                aliquota_iva = parse_decimal(row.get(colonne_trovate.get('aliquota_iva', ''), 0))
                if aliquota_iva == 0 and importo_iva > 0:
                    # Calcola aliquota dall'importo IVA
                    aliquota_iva = (importo_iva / importo_netto * 100) if importo_netto > 0 else Decimal('0')
                
                # Categoria e terreno
                categoria = parse_string(row.get(colonne_trovate.get('categoria', ''), ''))
                terreno_nome = parse_string(row.get(colonne_trovate.get('terreno', ''), ''))
                terreno_id = find_terreno(db, azienda_id, terreno_nome) if terreno_nome else None
                
                # Note
                note = parse_string(row.get(colonne_trovate.get('note', ''), ''))
                
                # Crea fattura (usando FatturaAmministrazione con tipo=entrata)
                fattura = FatturaAmministrazione(
                    azienda_id=azienda_id,
                    tipo=TipoFattura.ENTRATA,
                    numero=numero,
                    data_fattura=data_fattura,
                    data_registrazione=data_registrazione,
                    cliente_id=cliente_id,
                    cliente_nome=cliente_nome if not cliente_id else None,
                    cliente_piva=cliente_piva if not cliente_id else None,
                    cliente_cf=cliente_cf if not cliente_id else None,
                    importo_totale=importo_totale,
                    importo_iva=importo_iva,
                    importo_netto=importo_netto,
                    importo_incassato=importo_incassato,
                    stato_pagamento=stato_pagamento,
                    data_scadenza=data_scadenza,
                    data_incasso=data_incasso,
                    aliquota_iva=aliquota_iva,
                    categoria=categoria,
                    terreno_id=terreno_id,
                    note=note
                )
                
                db.add(fattura)
                importate += 1
                
            except Exception as e:
                errate += 1
                errori.append(f'Riga {idx + 2}: {str(e)}')
                continue
        
        # Commit
        db.commit()
        
        return {
            'success': True,
            'importate': importate,
            'errate': errate,
            'duplicate': duplicate,
            'errori': errori[:50]  # Limita a 50 errori
        }
        
    except Exception as e:
        db.rollback()
        return {
            'success': False,
            'error': str(e),
            'importate': 0,
            'errate': 0,
            'duplicate': 0
        }


def import_fatture_amministrazione(
    db: Session,
    file_path: str,
    skip_duplicates: bool = True
) -> Dict[str, any]:
    """
    Importa fatture amministrazione (ricevute) da file Excel
    
    Args:
        db: Database session
        file_path: Percorso del file Excel
        skip_duplicates: Se True, salta le fatture duplicate (stesso numero e data)
    
    Returns:
        Dict con statistiche dell'import
    """
    # Usa openpyxl direttamente invece di pandas per risparmiare ~300MB di RAM
    from openpyxl import load_workbook
    
    try:
        # Leggi il file Excel con openpyxl
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {'success': False, 'error': 'File Excel vuoto', 'importate': 0, 'errate': 0, 'duplicate': 0}
        
        # Prima riga = header, normalizza nomi colonne
        header = [str(col).strip().lower().replace(' ', '_') if col else f'col_{i}' for i, col in enumerate(rows[0])]
        data_rows = rows[1:]
        
        # Converti in lista di dizionari
        df_records = [dict(zip(header, row)) for row in data_rows]
        
        importate = 0
        errate = 0
        duplicate = 0
        errori = []
        
        # Mappa delle colonne possibili
        colonne_richieste = ['numero', 'data_fattura', 'importo_totale']
        colonne_opzionali = {
            'tipo': ['tipo', 'tipo_fattura', 'entrata_uscita'],
            'fornitore_nome': ['fornitore_nome', 'fornitore', 'nome_fornitore', 'ragione_sociale'],
            'fornitore_piva': ['fornitore_piva', 'piva', 'partita_iva', 'p.iva'],
            'fornitore_cf': ['fornitore_cf', 'cf', 'codice_fiscale'],
            'data_registrazione': ['data_registrazione', 'data_reg'],
            'importo_iva': ['importo_iva', 'iva'],
            'importo_netto': ['importo_netto', 'netto'],
            'importo_pagato': ['importo_pagato', 'pagato'],
            'stato_pagamento': ['stato_pagamento', 'stato'],
            'data_scadenza': ['data_scadenza', 'scadenza'],
            'data_pagamento': ['data_pagamento', 'data_pagamento_effettivo'],
            'periodo_da': ['periodo_da', 'periodo_inizio', 'da'],
            'periodo_a': ['periodo_a', 'periodo_fine', 'a'],
            'periodo_attribuzione': ['periodo_attribuzione', 'periodo'],
            'categoria': ['categoria', 'tipo_spesa'],
            'terreno': ['terreno', 'terreno_nome'],
            'note': ['note', 'osservazioni']
        }
        
        # Verifica colonne richieste
        colonne_trovate = {}
        for col in colonne_richieste:
            if col in header:
                colonne_trovate[col] = col
            else:
                return {
                    'success': False,
                    'error': f'Colonna obbligatoria "{col}" non trovata nel file',
                    'importate': 0,
                    'errate': 0,
                    'duplicate': 0
                }
        
        # Trova colonne opzionali
        for campo, possibili_nomi in colonne_opzionali.items():
            for nome in possibili_nomi:
                if nome in header:
                    colonne_trovate[campo] = nome
                    break
        
        # Processa ogni riga (df_records è lista di dict)
        for idx, row in enumerate(df_records):
            try:
                # Estrai dati base (row è un dict, accesso con .get())
                numero = parse_string(row.get(colonne_trovate['numero']))
                data_fattura = parse_date(row.get(colonne_trovate['data_fattura']))
                importo_totale = parse_decimal(row.get(colonne_trovate['importo_totale']))
                
                if not numero or not data_fattura or importo_totale == 0:
                    errate += 1
                    errori.append(f'Riga {idx + 2}: Dati obbligatori mancanti o non validi')
                    continue
                
                # Verifica duplicati
                if skip_duplicates:
                    esistente = db.query(FatturaAmministrazione).filter(
                        and_(
                            FatturaAmministrazione.numero == numero,
                            FatturaAmministrazione.data_fattura == data_fattura,
                            FatturaAmministrazione.deleted_at.is_(None)
                        )
                    ).first()
                    
                    if esistente:
                        duplicate += 1
                        continue
                
                # Tipo fattura
                tipo_str = parse_string(row.get(colonne_trovate.get('tipo', ''), ''))
                tipo = parse_tipo_fattura(tipo_str) if tipo_str else TipoFattura.USCITA
                
                # Fornitore
                fornitore_nome = parse_string(row.get(colonne_trovate.get('fornitore_nome', ''), ''))
                fornitore_piva = parse_string(row.get(colonne_trovate.get('fornitore_piva', ''), ''))
                fornitore_cf = parse_string(row.get(colonne_trovate.get('fornitore_cf', ''), ''))
                fornitore_id = find_fornitore(db, fornitore_nome, fornitore_piva, fornitore_cf)
                
                # Calcola importi
                importo_iva = parse_decimal(row.get(colonne_trovate.get('importo_iva', ''), 0))
                importo_netto = parse_decimal(row.get(colonne_trovate.get('importo_netto', ''), 0))
                if importo_netto == 0:
                    importo_netto = importo_totale - importo_iva
                
                importo_pagato = parse_decimal(row.get(colonne_trovate.get('importo_pagato', ''), 0))
                
                # Stato pagamento
                stato_str = parse_string(row.get(colonne_trovate.get('stato_pagamento', ''), ''))
                stato_pagamento = parse_stato_pagamento_amministrazione(stato_str) if stato_str else StatoPagamento.DA_PAGARE
                
                # Date
                data_registrazione = parse_date(row.get(colonne_trovate.get('data_registrazione', ''), '')) or date.today()
                data_scadenza = parse_date(row.get(colonne_trovate.get('data_scadenza', ''), ''))
                data_pagamento = parse_date(row.get(colonne_trovate.get('data_pagamento', ''), ''))
                
                # Periodo attribuzione
                periodo_da = parse_date(row.get(colonne_trovate.get('periodo_da', ''), ''))
                periodo_a = parse_date(row.get(colonne_trovate.get('periodo_a', ''), ''))
                periodo_attribuzione = parse_string(row.get(colonne_trovate.get('periodo_attribuzione', ''), ''))
                
                # Categoria e terreno (terreno non supportato per FatturaAmministrazione senza azienda_id)
                categoria = parse_string(row.get(colonne_trovate.get('categoria', ''), ''))
                
                # Note
                note = parse_string(row.get(colonne_trovate.get('note', ''), ''))
                
                # Crea fattura
                fattura = FatturaAmministrazione(
                    tipo=tipo,
                    numero=numero,
                    data_fattura=data_fattura,
                    data_registrazione=data_registrazione,
                    fornitore_id=fornitore_id,
                    importo_totale=importo_totale,
                    importo_iva=importo_iva,
                    importo_netto=importo_netto,
                    importo_pagato=importo_pagato,
                    stato_pagamento=stato_pagamento,
                    data_scadenza=data_scadenza,
                    data_pagamento=data_pagamento,
                    periodo_da=periodo_da,
                    periodo_a=periodo_a,
                    periodo_attribuzione=periodo_attribuzione,
                    categoria=categoria,
                    note=note
                )
                
                db.add(fattura)
                importate += 1
                
            except Exception as e:
                errate += 1
                errori.append(f'Riga {idx + 2}: {str(e)}')
                continue
        
        # Commit
        db.commit()
        
        return {
            'success': True,
            'importate': importate,
            'errate': errate,
            'duplicate': duplicate,
            'errori': errori[:50]  # Limita a 50 errori
        }
        
    except Exception as e:
        db.rollback()
        return {
            'success': False,
            'error': str(e),
            'importate': 0,
            'errate': 0,
            'duplicate': 0
        }

